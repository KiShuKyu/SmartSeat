from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
import sqlite3
import datetime
import os
import openpyxl
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="SmartSeat", version="2.0.0")

DB_PATH = "smartseat.db"
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN")

# Database

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute('''
        CREATE TABLE IF NOT EXISTS seats (
            id            TEXT PRIMARY KEY,
            status        TEXT DEFAULT 'vacant',
            student_name  TEXT,
            urn           INTEGER,
            reserved_until TIMESTAMP
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS borrows (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            urn          INTEGER NOT NULL,
            student_name TEXT    NOT NULL,
            book_title   TEXT    NOT NULL,
            author       TEXT,
            borrow_time  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            returned     BOOLEAN   DEFAULT 0,
            return_time  TIMESTAMP
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS entries (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            urn        INTEGER NOT NULL,
            entry_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS books_inventory (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            title     TEXT NOT NULL,
            author    TEXT,
            shelf     TEXT NOT NULL,
            quantity  INTEGER DEFAULT 1,
            available INTEGER DEFAULT 1
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS book_loans (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            urn            INTEGER NOT NULL,
            student_name   TEXT NOT NULL,
            book_title     TEXT NOT NULL,
            borrow_date    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            due_date       TIMESTAMP,
            return_date    TIMESTAMP,
            penalty_amount INTEGER DEFAULT 0,
            returned       BOOLEAN DEFAULT 0
        )
    ''')

    # Seed seats if empty
    cur.execute("SELECT COUNT(*) FROM seats")
    if cur.fetchone()[0] == 0:
        seats = [f"T{i}" for i in range(1, 13)]
        for table in ['A', 'B']:
            for row in range(1, 7):
                seats += [f"{table}{row}L", f"{table}{row}R"]
        cur.executemany(
            "INSERT INTO seats (id, status) VALUES (?, 'vacant')",
            [(s,) for s in seats]
        )

    conn.commit()
    conn.close()

# Helpers

def _expire_seats(cur):
    """Auto-vacate seats whose reservation time has passed."""
    now = datetime.datetime.now().isoformat()
    cur.execute("""
        UPDATE seats
        SET status='vacant', student_name=NULL, urn=NULL, reserved_until=NULL
        WHERE status='occupied' AND reserved_until < ?
    """, (now,))

# Request Models

class ReserveRequest(BaseModel):
    seat_id: str
    urn: int
    student_name: str
    duration_hours: int = 2

class ReleaseRequest(BaseModel):
    seat_id: str
    urn: int

class BorrowRequest(BaseModel):
    urn: int
    student_name: str
    book_title: str
    author: Optional[str] = None

class ReturnRequest(BaseModel):
    record_id: int
class AdminLoginRequest(BaseModel):
    password: str

class BorrowBookRequest(BaseModel):
    urn: int
    student_name: str
    book_title: str

class ReturnBookRequest(BaseModel):
    loan_id: int

# API Endpoints

@app.get("/api/seats")
async def get_seats():
    conn = get_conn()
    cur = conn.cursor()
    _expire_seats(cur)
    conn.commit()

    cur.execute("SELECT * FROM seats ORDER BY id")
    rows = cur.fetchall()
    conn.close()

    return [
        {
            "id":           r["id"],
            "status":       r["status"],
            "student_name": r["student_name"],
            "urn":          r["urn"],
            "end_time":     r["reserved_until"],
        }
        for r in rows
    ]


@app.post("/api/reserve")
async def reserve_seat(req: ReserveRequest):
    # Validate inputs
    if not req.seat_id or not req.student_name.strip():
        raise HTTPException(400, "seat_id and student_name are required")
    if req.urn <= 0:
        raise HTTPException(400, "URN must be a positive integer")

    duration = max(1, min(req.duration_hours, 3))   # clamp 1-3 hrs

    conn = get_conn()
    cur = conn.cursor()
    _expire_seats(cur)

    # Check seat exists and is vacant
    cur.execute("SELECT status FROM seats WHERE id=?", (req.seat_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Seat '{req.seat_id}' does not exist")
    if row["status"] != "vacant":
        conn.close()
        raise HTTPException(400, "Seat is not available right now")

    # One active reservation per student
    cur.execute("""
        SELECT id FROM seats
        WHERE urn=? AND status='occupied' AND reserved_until > ?
    """, (req.urn, datetime.datetime.now().isoformat()))
    if cur.fetchone():
        conn.close()
        raise HTTPException(400, "You already have an active seat reservation")

    # Log entry
    cur.execute("INSERT INTO entries (urn) VALUES (?)", (req.urn,))

    # Reserve
    end_time = datetime.datetime.now() + datetime.timedelta(hours=duration)
    cur.execute("""
        UPDATE seats
        SET status='occupied', student_name=?, urn=?, reserved_until=?
        WHERE id=?
    """, (req.student_name.strip(), req.urn, end_time.isoformat(), req.seat_id))

    conn.commit()
    conn.close()
    return {"message": "Seat reserved successfully", "reserved_until": end_time.isoformat()}


@app.post("/api/release")
async def release_seat(req: ReleaseRequest):
    if req.urn <= 0:
        raise HTTPException(400, "URN must be a positive integer")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT urn, status FROM seats WHERE id=?", (req.seat_id,))
    row = cur.fetchone()

    if not row:
        conn.close()
        raise HTTPException(404, f"Seat '{req.seat_id}' does not exist")
    if row["status"] != "occupied":
        conn.close()
        raise HTTPException(400, "Seat is not currently reserved")
    if row["urn"] != req.urn:
        conn.close()
        raise HTTPException(403, "URN does not match - you can only release your own seat")

    cur.execute("""
        UPDATE seats
        SET status='vacant', student_name=NULL, urn=NULL, reserved_until=NULL
        WHERE id=?
    """, (req.seat_id,))
    conn.commit()
    conn.close()
    return {"message": "Seat released successfully"}


@app.get("/api/stats")
async def get_stats():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM seats WHERE status='vacant'")
    vacant = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM seats WHERE status='occupied'")
    occupied = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM borrows WHERE returned=0")
    active_borrows = cur.fetchone()[0]
    today = datetime.date.today().isoformat()
    cur.execute("SELECT COUNT(*) FROM entries WHERE date(entry_time)=?", (today,))
    today_entries = cur.fetchone()[0]

    conn.close()
    return {
        "vacant":        vacant,
        "occupied":      occupied,
        "active_borrows": active_borrows,
        "today_entries": today_entries,
    }


@app.get("/api/reservations")
async def get_reservations():
    conn = get_conn()
    cur = conn.cursor()
    _expire_seats(cur)
    conn.commit()

    cur.execute("""
        SELECT id AS seat_id, student_name, urn, reserved_until AS end_time
        FROM seats
        WHERE status='occupied'
        ORDER BY reserved_until
    """)
    rows = cur.fetchall()
    conn.close()

    now = datetime.datetime.now()
    return [
        {
            "seat_id":      r["seat_id"],
            "urn":          r["urn"],
            "student_name": r["student_name"],
            "end_time":     r["end_time"],
            "active":       datetime.datetime.fromisoformat(r["end_time"]) > now
                            if r["end_time"] else False,
        }
        for r in rows
    ]


@app.get("/api/borrows")
async def get_borrows():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM borrows ORDER BY borrow_time DESC")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/borrow")
async def borrow_book(req: BorrowRequest):
    if req.urn <= 0:
        raise HTTPException(400, "URN must be a positive integer")
    if not req.student_name.strip():
        raise HTTPException(400, "student_name is required")
    if not req.book_title.strip():
        raise HTTPException(400, "book_title is required")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO borrows (urn, student_name, book_title, author)
        VALUES (?, ?, ?, ?)
    """, (req.urn, req.student_name.strip(), req.book_title.strip(), req.author))
    conn.commit()
    conn.close()
    return {"message": "Book borrowed successfully"}


@app.post("/api/return")
async def return_book(req: ReturnRequest):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM borrows WHERE id=? AND returned=0",
        (req.record_id,)
    )
    borrow = cur.fetchone()
    if not borrow:
        conn.close()
        raise HTTPException(400, "Record not found or already returned")

    cur.execute("""
        UPDATE borrows
        SET returned=1, return_time=CURRENT_TIMESTAMP
        WHERE id=?
    """, (req.record_id,))

    # Also mark matching book_loans as returned and restore inventory
    cur.execute("""
        UPDATE book_loans
        SET returned=1, return_date=?, penalty_amount=0
        WHERE urn=? AND book_title=? AND returned=0
        ORDER BY borrow_date DESC LIMIT 1
    """, (datetime.datetime.now().isoformat(), borrow["urn"], borrow["book_title"]))

    cur.execute(
        "UPDATE books_inventory SET available = available + 1 WHERE title=?",
        (borrow["book_title"],)
    )

    conn.commit()
    conn.close()
    return {"message": "Book returned successfully"}


@app.post("/api/entry")
async def log_entry(data: dict):
    urn = data.get("urn")
    if not urn:
        raise HTTPException(400, "urn is required")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO entries (urn) VALUES (?)", (urn,))
    conn.commit()
    conn.close()
    return {"message": "Entry logged"}


# ─── Book Inventory & Loan Endpoints ───

@app.get("/api/books")
async def get_books_inventory():
    """Return all books from inventory."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM books_inventory ORDER BY title")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/borrow-book")
async def borrow_book_from_inventory(req: BorrowBookRequest):
    """Borrow a book from inventory with penalty tracking."""
    if req.urn <= 0:
        raise HTTPException(400, "URN must be a positive integer")
    if not req.student_name.strip():
        raise HTTPException(400, "student_name is required")
    if not req.book_title.strip():
        raise HTTPException(400, "book_title is required")

    conn = get_conn()
    cur = conn.cursor()

    # Check if student already has 3+ active loans
    cur.execute(
        "SELECT COUNT(*) FROM book_loans WHERE urn=? AND returned=0",
        (req.urn,)
    )
    active = cur.fetchone()[0]
    if active >= 3:
        conn.close()
        raise HTTPException(400, f"Student already has {active} active loans (max 3)")

    # Check book availability
    cur.execute(
        "SELECT id, available FROM books_inventory WHERE title=?",
        (req.book_title.strip(),)
    )
    book = cur.fetchone()
    if not book:
        conn.close()
        raise HTTPException(404, f"Book '{req.book_title}' not found in inventory")
    if book["available"] <= 0:
        conn.close()
        raise HTTPException(400, f"No copies of '{req.book_title}' available")

    # Create loan
    now = datetime.datetime.now()
    due = now + datetime.timedelta(days=15)
    cur.execute("""
        INSERT INTO book_loans (urn, student_name, book_title, borrow_date, due_date)
        VALUES (?, ?, ?, ?, ?)
    """, (req.urn, req.student_name.strip(), req.book_title.strip(),
          now.isoformat(), due.isoformat()))

    # Decrease available quantity
    cur.execute(
        "UPDATE books_inventory SET available = available - 1 WHERE id=?",
        (book["id"],)
    )

    # Also log legacy borrow
    cur.execute("""
        INSERT INTO borrows (urn, student_name, book_title)
        VALUES (?, ?, ?)
    """, (req.urn, req.student_name.strip(), req.book_title.strip()))

    conn.commit()
    conn.close()
    return {"message": "Book borrowed successfully", "due_date": due.isoformat()}

@app.post("/api/return-book")
async def return_book_with_penalty(req: ReturnBookRequest):
    """Return a borrowed book, calculating penalty if late."""
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM book_loans WHERE id=? AND returned=0",
        (req.loan_id,)
    )
    loan = cur.fetchone()
    if not loan:
        conn.close()
        raise HTTPException(400, "Loan not found or already returned")

    now = datetime.datetime.now()
    due = datetime.datetime.fromisoformat(loan["due_date"])
    days_late = max(0, (now - due).days)
    penalty = days_late * 50

    cur.execute("""
        UPDATE book_loans
        SET returned=1, return_date=?, penalty_amount=?
        WHERE id=?
    """, (now.isoformat(), penalty, req.loan_id))

    # Increase available quantity
    cur.execute(
        "UPDATE books_inventory SET available = available + 1 WHERE title=?",
        (loan["book_title"],)
    )

    # Mark legacy borrow as returned too
    cur.execute("""
        UPDATE borrows SET returned=1, return_time=CURRENT_TIMESTAMP
        WHERE urn=? AND book_title=? AND returned=0
        ORDER BY borrow_time DESC LIMIT 1
    """, (loan["urn"], loan["book_title"]))

    conn.commit()
    conn.close()
    return {
        "message": "Book returned successfully",
        "days_late": days_late,
        "penalty_amount": penalty,
    }

@app.post("/api/admin/upload-books")
async def upload_books_excel(request: Request, file: UploadFile = File(...)):
    """Upload Excel file with book inventory."""
    _require_admin(request)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active

        # Look for header row
        headers = [str(c.value).strip().lower() for c in ws[1]]
        required = {'title', 'author', 'shelf', 'quantity'}
        if not required.issubset(set(headers)):
            raise HTTPException(
                400,
                f"Excel must have columns: title, author, shelf, quantity (found: {headers})"
            )

        title_idx = headers.index('title')
        author_idx = headers.index('author')
        shelf_idx = headers.index('shelf')
        qty_idx = headers.index('quantity')

        conn = get_conn()
        cur = conn.cursor()
        inserted = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            title = str(row[title_idx]).strip() if row[title_idx] else None
            if not title:
                continue
            author = str(row[author_idx]).strip() if row[author_idx] else None
            shelf = str(row[shelf_idx]).strip() if row[shelf_idx] else "Unshelved"
            try:
                qty = int(row[qty_idx]) if row[qty_idx] else 1
            except (ValueError, TypeError):
                qty = 1

            cur.execute(
                "INSERT INTO books_inventory (title, author, shelf, quantity, available) VALUES (?, ?, ?, ?, ?)",
                (title, author, shelf, qty, qty)
            )
            inserted += 1

        conn.commit()
        conn.close()
        return {"message": f"Inserted {inserted} books into inventory"}
    except HTTPException:
        raise
    except Exception as e:
        msg = str(e)
        if "is not a zip file" in msg or "BadZipFile" in msg:
            raise HTTPException(400, "The uploaded file is not a valid .xlsx file. This error occurs when uploading old .xls (binary) files or corrupted files — please save it as .xlsx first.")
        raise HTTPException(400, f"Failed to parse Excel file: {msg}")

@app.get("/api/admin/loans")
async def get_active_loans(request: Request):
    """Return all book loans (active and returned)."""
    _require_admin(request)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM book_loans ORDER BY borrow_date DESC")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


# Admin Endpoints
@app.post("/api/admin/clear-expired")
async def clear_expired_reservations(request: Request):
    _require_admin(request)
    conn = get_conn()
    cur = conn.cursor()
    _expire_seats(cur)
    cleared = cur.rowcount
    conn.commit()
    conn.close()
    return {"message": f"Cleared {cleared} expired seat reservations"}


@app.post("/api/admin/reset-demo")
async def reset_demo_data(request: Request):
    _require_admin(request)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM borrows")
    cur.execute("DELETE FROM book_loans")
    cur.execute("DELETE FROM books_inventory")
    cur.execute("DELETE FROM entries")
    cur.execute("""
        UPDATE seats
        SET status='vacant', student_name=NULL, urn=NULL, reserved_until=NULL
    """)
    conn.commit()
    conn.close()
    return {"message": "All data reset successfully"}

@app.post("/api/admin/clear-returned")
async def clear_returned_books(request: Request):
    _require_admin(request)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM borrows WHERE returned=1")
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    return {"message": f"Cleared {deleted} returned book records"}

@app.post("/api/admin/clear-returned-loans")
async def clear_returned_loans(request: Request):
    _require_admin(request)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM book_loans WHERE returned=1")
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    return {"message": f"Cleared {deleted} returned loan records"}

@app.post("/api/admin/login")
async def admin_login(req: AdminLoginRequest):
    if req.password != ADMIN_PASSWORD:
        raise HTTPException(401, "Invalid admin credentials")
    return {"token": ADMIN_TOKEN}


# Static files & root

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def root():
    return FileResponse("static/index.html")


@app.on_event("startup")
async def startup_event():
    init_db()

def _require_admin(request: Request):
    auth = request.headers.get("authorization") or request.headers.get("Authorization")
    if not auth or not auth.startswith("Bearer "):
        raise HTTPException(401, "Admin authorization required")
    token = auth.split(" ", 1)[1]
    if token != ADMIN_TOKEN:
        raise HTTPException(403, "Invalid admin token")
